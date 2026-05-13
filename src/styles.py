"""Reusable openpyxl style objects."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_FONT = "Proxima Nova"

# Full ARGB hex strings (FF prefix = fully opaque)
_NAVY = "FF152534"
_PINK = "FFE21A6B"
_BLUE = "FF1F4E79"
_WHITE = "FFFFFFFF"
_GRAY = "FF888888"

# Header fill colors — SA uses a slightly different pink to match reference exactly
_HEADER_PINK_SA = "FFE2216B"
_HEADER_PINK = "FFE21A6B"

# --- Title row fonts (no fill — colored text on plain background) ---
TITLE_FONT = Font(name=_FONT, size=15, bold=True, color=_NAVY)       # SA
TITLE_FONT_SC = Font(name=_FONT, size=12, bold=True, color=_NAVY)    # SC
TITLE_FONT_APR = Font(name=_FONT, size=12, bold=True, color=_PINK)   # APR
TITLE_FONT_BM = Font(name=_FONT, size=15, bold=True, color=_BLUE)    # BM

# --- Section/subsection row fonts (no fill — colored text) ---
SECTION_FONT = Font(name=_FONT, size=12, bold=True, color=_NAVY)         # SA/SC sections
SECTION_FONT_SA_SUB = Font(name=_FONT, size=14, bold=True, color=_NAVY)  # SA "Client Sources"
SECTION_FONT_APR = Font(name=_FONT, size=12, bold=True, color=_PINK)     # APR sections
SECTION_FONT_BM = Font(name=_FONT, size=12, bold=True, color=_BLUE)      # BM sections

# --- Column header row fonts (pink fill, white text) ---
HEADER_FONT = Font(name=_FONT, size=12, bold=True, color=_WHITE)

# --- Data / utility fonts ---
DATA_FONT = Font(name=_FONT, size=12)
DATA_FONT_BOLD = Font(name=_FONT, size=12, bold=True)
EMPTY_SECTION_FONT = Font(name=_FONT, size=12, italic=True, color=_GRAY)

# --- Fills ---
HEADER_FILL = PatternFill("solid", fgColor=_HEADER_PINK)       # APR, SC, BM header rows
HEADER_FILL_SA = PatternFill("solid", fgColor=_HEADER_PINK_SA) # SA header row (slightly different pink)

# --- Alignment ---
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(horizontal="center", vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
CENTER = Alignment(horizontal="center")

# --- Borders ---
THIN_GRAY = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
