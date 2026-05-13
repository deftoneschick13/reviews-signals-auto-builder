"""Generates the output Reviews Signals .xlsx workbook from analyzer results."""

import logging
from pathlib import Path

from openpyxl import Workbook

log = logging.getLogger(__name__)
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from src.analyzers.ai_platform_response import (
    CATEGORIES_ORDERED,
    PlatformResponseRow,
    build_ai_platform_response,
)
from src.analyzers.benchmarking import BenchmarkRow, build_benchmarking
from src.analyzers.sentiment_cooccurrence import (
    CoOccurrenceRow,
    DetailedSentimentRow,
    SummaryRow,
    build_sentiment_cooccurrence,
)
from src.analyzers.source_attribution import SourceRow, build_source_attribution
from src.matchers import LabeledChat
from src.prompt_library import PromptEntry
from src.styles import (
    DATA_FONT,
    DATA_FONT_BOLD,
    EMPTY_SECTION_FONT,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FILL_SA,
    HEADER_FONT,
    SECTION_FONT,
    SECTION_FONT_APR,
    SECTION_FONT_BM,
    SECTION_FONT_SA_SUB,
    THIN_GRAY,
    TITLE_FONT,
    TITLE_FONT_APR,
    TITLE_FONT_BM,
    TITLE_FONT_SC,
    WRAP_TOP,
)

_APR_HEADERS = [
    "Prompt ID", "Prompt", "Brand Mentioned?", "Position",
    "Context Analysis", "Sentiment Score", "Sentiment",
    "Co-Mentions", "Sources/Citations", "Chat Snapshot", "Notes",
]
# Reference column widths (matched from Babylon Tours reference workbook)
_APR_COL_WIDTHS = {
    1: 18.63, 2: 45.0, 3: 14.0, 4: 10.0, 5: 70.0, 6: 14.0,
    7: 14.0, 8: 35.0, 9: 50.0, 10: 60.0, 11: 50.0,
}
_APR_COLS = len(_APR_HEADERS)

_APR_DATA_ROW_HEIGHT = 112.5


def build_workbook(
    chats: list[LabeledChat],
    prompt_library: dict[str, PromptEntry],
    brand_name: str,
    date_range_str: str,
    output_path: Path | str,
) -> Path:
    """Create the output workbook at output_path. Returns the path."""
    wb = Workbook()
    wb.remove(wb.active)

    sa_rows = build_source_attribution(chats)
    sa_ws = wb.create_sheet("Source Attribution Tracking")
    _build_source_attribution_sheet(sa_ws, sa_rows, brand_name, date_range_str)

    apr_data = build_ai_platform_response(chats, prompt_library, brand_name)
    apr_ws = wb.create_sheet("AI Platform Response Tracking")
    _build_ai_platform_response_sheet(apr_ws, apr_data, brand_name, date_range_str)

    sc_summary, sc_coocc, sc_detailed = build_sentiment_cooccurrence(chats, prompt_library, brand_name)
    sc_ws = wb.create_sheet("Sentiment & Co-Occurrence")
    _build_sentiment_cooccurrence_sheet(sc_ws, sc_summary, sc_coocc, sc_detailed, brand_name, date_range_str)

    bm_data = build_benchmarking(chats, prompt_library, brand_name)
    bm_ws = wb.create_sheet("Benchmarking")
    _build_benchmarking_sheet(bm_ws, bm_data, brand_name, date_range_str)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("build_workbook: saved %d chats → %s", len(chats), output_path)
    return output_path


def _merge_write(
    ws: Worksheet,
    row: int,
    n_cols: int,
    value,
    font,
    fill=None,
    height=None,
    alignment: Alignment | None = None,
) -> None:
    end_col_letter = chr(ord("A") + n_cols - 1)
    ws.merge_cells(f"A{row}:{end_col_letter}{row}")
    c = ws[f"A{row}"]
    c.value = value
    c.font = font
    if fill:
        c.fill = fill
    if height:
        ws.row_dimensions[row].height = height
    if alignment:
        c.alignment = alignment


def _write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN


def _build_source_attribution_sheet(
    ws: Worksheet,
    rows: list[SourceRow],
    brand_name: str,
    date_range_str: str,
) -> None:
    # R1: title (no fill, dark navy text, centered)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Source Attribution Tracking — {brand_name} — {date_range_str}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # R2: "Client Sources" section (no fill, dark navy, size 14)
    ws.merge_cells("A2:F2")
    sub_cell = ws["A2"]
    sub_cell.value = "Client Sources"
    sub_cell.font = SECTION_FONT_SA_SUB

    # R3: column headers (SA pink fill, white text, centered)
    headers = [
        "Domain", "Source URL", "Content Type", "Topic",
        "Platform Citations", "Citation Count",
    ]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL_SA
        c.alignment = HEADER_ALIGN

    # R4+: data
    if not rows:
        msg = ws.cell(row=4, column=1, value="No source data in the selected date range.")
        msg.font = EMPTY_SECTION_FONT
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
    else:
        for row_idx, sr in enumerate(rows, start=4):
            values = [
                sr.domain, sr.source_url, sr.content_type,
                sr.topic, sr.platform_citations, sr.citation_count,
            ]
            for col_idx, v in enumerate(values, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=v)
                c.font = DATA_FONT
                c.alignment = WRAP_TOP
                c.border = THIN_GRAY

    # Column widths matched to reference
    widths = {1: 48.5, 2: 70.0, 3: 61.63, 4: 55.0, 5: 20.63, 6: 16.0}
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = width


def _build_ai_platform_response_sheet(
    ws: Worksheet,
    data: dict,
    brand_name: str,
    date_range_str: str,
) -> None:
    n = _APR_COLS
    current_row = 1

    # R1: title bar (pink text, no fill, centered)
    _merge_write(
        ws, current_row, n,
        f"AI Platform Response Tracking — {brand_name} — {date_range_str}",
        TITLE_FONT_APR,
        alignment=Alignment(horizontal="center"),
    )
    current_row += 1

    # R2: blank
    current_row += 1

    for platform in sorted(data.keys()):
        # Platform section header (pink text, no fill)
        _merge_write(ws, current_row, n, f"Platform: {platform}", SECTION_FONT_APR)
        current_row += 1

        for category in CATEGORIES_ORDERED:
            # Category sub-section header (pink text, no fill)
            _merge_write(
                ws, current_row, n,
                f"Platform: {platform} | {category}",
                SECTION_FONT_APR,
            )
            current_row += 1

            # Column headers (pink fill, white text, centered)
            _write_header_row(ws, current_row, _APR_HEADERS)
            current_row += 1

            rows: list[PlatformResponseRow] = data[platform].get(category, [])
            if not rows:
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=n,
                )
                msg = ws.cell(
                    row=current_row, column=1,
                    value="No data for this category in the selected date range.",
                )
                msg.font = EMPTY_SECTION_FONT
                current_row += 1
            else:
                for pr in rows:
                    values = [
                        pr.prompt_id, pr.prompt, pr.brand_mentioned, pr.position,
                        pr.context_analysis, pr.sentiment_score, pr.sentiment_label,
                        pr.co_mentions, pr.sources_citations, pr.chat_snapshot, pr.notes,
                    ]
                    for col_idx, v in enumerate(values, start=1):
                        c = ws.cell(row=current_row, column=col_idx, value=v)
                        c.font = DATA_FONT
                        c.alignment = WRAP_TOP
                        c.border = THIN_GRAY
                    ws.row_dimensions[current_row].height = _APR_DATA_ROW_HEIGHT
                    current_row += 1

            # Blank row separator
            current_row += 1

    # Column widths matched to reference
    for col_idx, width in _APR_COL_WIDTHS.items():
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = width

    ws.freeze_panes = "A3"


_BM_N_COLS = 5
# Column widths matched to reference
_BM_COL_WIDTHS = {1: 18.0, 2: 19.75, 3: 10.5, 4: 18.38, 5: 105.75}
_BM_HEADERS = ["Brand", "Mention Rate", "Avg. Position", "Avg. Sentiment", "Dominant Themes"]


def _build_benchmarking_sheet(
    ws: Worksheet,
    data: dict,
    brand_name: str,
    date_range_str: str,
) -> None:
    n = _BM_N_COLS
    r = 1

    # R1: title (deep blue text, no fill, centered)
    _merge_write(
        ws, r, n,
        f"Benchmarking — {brand_name} — {date_range_str}",
        TITLE_FONT_BM,
        alignment=Alignment(horizontal="center"),
    )
    r += 1

    r += 1  # blank

    for category in CATEGORIES_ORDERED:
        # Section header (deep blue text, no fill)
        _merge_write(ws, r, n, category, SECTION_FONT_BM)
        r += 1

        # Column headers (pink fill, white text, centered)
        _write_header_row(ws, r, _BM_HEADERS)
        r += 1

        rows: list[BenchmarkRow] = data.get(category, [])
        for i, br in enumerate(rows):
            values = [br.brand, br.mention_rate, br.avg_position, br.avg_sentiment, br.dominant_themes]
            for col_idx, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=col_idx, value=v)
                c.font = DATA_FONT
                c.alignment = WRAP_TOP
                c.border = THIN_GRAY
            # Focal brand (first row per section) bolded
            if i == 0:
                ws.cell(row=r, column=1).font = DATA_FONT_BOLD
            r += 1

        r += 1  # blank separator

    for col_idx, width in _BM_COL_WIDTHS.items():
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = width


_SC_N_COLS = 8
# Column widths matched to reference
_SC_COL_WIDTHS = {1: 19.13, 2: 21.63, 3: 22.0, 4: 25.75, 5: 16.0, 7: 25.88, 8: 19.5}

_SC_SUMMARY_HEADERS_TPL = [
    "Category", "Total Prompts", "{brand} Mentioned", "Mention Rate",
    "Avg. Sentiment Score", "Positive", "Neutral", "Negative",
]
_SC_COOCC_HEADERS = [
    "Brand/Entity", "Co-Occurrence Count", "Relationship Type",
    "Typical Position vs {brand}", "Key Associations", "Opportunity/Threat",
]
_SC_DETAIL_HEADERS_TPL = [
    "Prompt ID", "Prompt", "Category", "{brand} Mentioned",
    "Sentiment Score", "Sentiment Label", "Key Observations",
]


def _build_sentiment_cooccurrence_sheet(
    ws: Worksheet,
    summary_rows: list[SummaryRow],
    coocc_rows: list[CoOccurrenceRow],
    detailed_rows: list[DetailedSentimentRow],
    brand_name: str,
    date_range_str: str,
) -> None:
    n = _SC_N_COLS
    r = 1

    # R1: title (dark navy text, no fill)
    _merge_write(
        ws, r, n,
        f"Sentiment & Co-Occurrence — {brand_name} — {date_range_str}",
        TITLE_FONT_SC,
    )
    r += 1

    # R2: blank
    r += 1

    # --- Section 1: Sentiment Analysis Summary ---
    _merge_write(ws, r, n, "Sentiment Analysis Summary", SECTION_FONT)
    r += 1
    summary_headers = [h.replace("{brand}", brand_name) for h in _SC_SUMMARY_HEADERS_TPL]
    _write_header_row(ws, r, summary_headers)
    r += 1
    for sr in summary_rows:
        values = [
            sr.category, sr.total_prompts, sr.brand_mentioned_count,
            sr.mention_rate, sr.avg_sentiment_score,
            sr.positive_count, sr.neutral_count, sr.negative_count,
        ]
        is_overall = sr.category == "OVERALL"
        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx, value=v)
            c.font = DATA_FONT_BOLD if is_overall else DATA_FONT
            c.alignment = WRAP_TOP
            c.border = THIN_GRAY
        r += 1

    r += 2  # 2 blank rows

    # --- Section 2: Co-Occurrence Analysis ---
    _merge_write(ws, r, n, "Co-Occurrence Analysis", SECTION_FONT)
    r += 1
    coocc_headers = [h.replace("{brand}", brand_name) for h in _SC_COOCC_HEADERS]
    _write_header_row(ws, r, coocc_headers)
    r += 1
    if not coocc_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        msg = ws.cell(row=r, column=1, value="No co-occurrence data in the selected date range.")
        msg.font = EMPTY_SECTION_FONT
        r += 1
    else:
        for cr in coocc_rows:
            values = [
                cr.brand_or_entity, cr.cooccurrence_count, cr.relationship_type,
                cr.typical_position, cr.key_associations, cr.opportunity_threat,
            ]
            for col_idx, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=col_idx, value=v)
                c.font = DATA_FONT
                c.alignment = WRAP_TOP
                c.border = THIN_GRAY
            r += 1

    r += 2  # 2 blank rows

    # --- Section 3: Detailed Sentiment by Prompt ---
    _merge_write(ws, r, n, "Detailed Sentiment by Prompt", SECTION_FONT)
    r += 1
    detail_headers = [h.replace("{brand}", brand_name) for h in _SC_DETAIL_HEADERS_TPL]
    _write_header_row(ws, r, detail_headers)
    r += 1
    if not detailed_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        msg = ws.cell(row=r, column=1, value="No detailed sentiment data in the selected date range.")
        msg.font = EMPTY_SECTION_FONT
        r += 1
    else:
        for dr in detailed_rows:
            values = [
                dr.prompt_id, dr.prompt, dr.category, dr.brand_mentioned,
                dr.sentiment_score, dr.sentiment_label, dr.key_observations,
            ]
            for col_idx, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=col_idx, value=v)
                c.font = DATA_FONT
                c.alignment = WRAP_TOP
                c.border = THIN_GRAY
            r += 1

    for col_idx, width in _SC_COL_WIDTHS.items():
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = width
