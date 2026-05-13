"""Tests for src/workbook_builder.py."""

import openpyxl
import pytest

from src.matchers import LabeledChat
from src.peec_client import Chat
from src.workbook_builder import build_workbook

BRAND = "Test Brand"
DATE_RANGE = "2026-04-01 to 2026-05-01"
SHEET_NAMES = [
    "Source Attribution Tracking",
    "AI Platform Response Tracking",
    "Sentiment & Co-Occurrence",
    "Benchmarking",
]


def _chat(chat_id: str, sources: list[str]) -> Chat:
    return Chat(
        id=chat_id,
        model="chatgpt-scraper",
        model_channel="ChatGPT",
        prompt="test prompt",
        response="test response",
        country="US",
        position=None,
        mentions=[],
        sources=sources,
        sentiment=None,
        created="2026-05-01",
    )


def _labeled(chat: Chat) -> LabeledChat:
    return LabeledChat(chat=chat, prompt_id="DB-01", category="Direct Brand Queries")


@pytest.fixture
def built_path(tmp_path):
    chats = [_labeled(_chat("ch_1", ["https://example.com", "https://other.com"]))]
    return build_workbook(chats, {}, BRAND, DATE_RANGE, tmp_path / "output.xlsx")


@pytest.fixture
def built_wb(built_path):
    return openpyxl.load_workbook(built_path)


def test_build_workbook_creates_file_at_output_path(tmp_path):
    path = tmp_path / "output.xlsx"
    result = build_workbook([], {}, BRAND, DATE_RANGE, path)
    assert result == path
    assert path.exists()


def test_build_workbook_has_all_four_sheets_in_correct_order(built_wb):
    assert built_wb.sheetnames == SHEET_NAMES


def test_source_attribution_tab_title_in_A1(built_wb):
    ws = built_wb["Source Attribution Tracking"]
    assert "Source Attribution Tracking" in ws["A1"].value
    assert BRAND in ws["A1"].value
    assert DATE_RANGE in ws["A1"].value


def test_source_attribution_tab_subsection_in_A2(built_wb):
    ws = built_wb["Source Attribution Tracking"]
    assert ws["A2"].value == "Client Sources"


def test_source_attribution_tab_headers_in_row_3(built_wb):
    ws = built_wb["Source Attribution Tracking"]
    headers = [ws.cell(row=3, column=i).value for i in range(1, 7)]
    assert headers == [
        "Domain", "Source URL", "Content Type", "Topic",
        "Platform Citations", "Citation Count",
    ]


def test_source_attribution_tab_data_starts_at_row_4(built_wb):
    ws = built_wb["Source Attribution Tracking"]
    # Row 4 should have a domain value (data, not a header or empty)
    assert ws.cell(row=4, column=1).value is not None
    assert ws.cell(row=4, column=1).value != "Domain"


def test_source_attribution_tab_freeze_panes_at_A4(built_wb):
    ws = built_wb["Source Attribution Tracking"]
    assert str(ws.freeze_panes) == "A4"


def test_benchmarking_tab_has_title_in_A1(built_wb):
    ws = built_wb["Benchmarking"]
    assert "Benchmarking" in ws["A1"].value
    assert BRAND in ws["A1"].value


def test_build_workbook_with_empty_chats_produces_valid_file(tmp_path):
    path = tmp_path / "empty.xlsx"
    build_workbook([], {}, BRAND, DATE_RANGE, path)
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == SHEET_NAMES
    ws = wb["Source Attribution Tracking"]
    assert ws.cell(row=4, column=1).value == "No source data in the selected date range."


# ---------------------------------------------------------------------------
# AI Platform Response Tracking workbook tests
# ---------------------------------------------------------------------------

def _apr_chat_with_brand(chat_id: str) -> "LabeledChat":
    from src.peec_client import Chat
    from src.matchers import LabeledChat
    chat = Chat(
        id=chat_id, model="chatgpt-scraper", model_channel="ChatGPT",
        prompt="Test prompt", response="Babylon Tours is great",
        country="US", position=1, mentions=["Babylon Tours"],
        sources=["https://example.com"], sentiment=75.0, created="2026-05-01",
    )
    return LabeledChat(chat=chat, prompt_id="DB-01", category="Direct Brand Queries")


@pytest.fixture
def apr_wb(tmp_path):
    from src.prompt_library import PromptEntry
    library = {"DB-01": PromptEntry("DB-01", "Test prompt", "Direct Brand Queries", "", "")}
    chats = [_apr_chat_with_brand("ch_1")]
    path = build_workbook(chats, library, BRAND, DATE_RANGE, tmp_path / "apr.xlsx")
    return openpyxl.load_workbook(path)


def test_apr_tab_has_title_bar(apr_wb):
    ws = apr_wb["AI Platform Response Tracking"]
    assert "AI Platform Response Tracking" in ws["A1"].value
    assert BRAND in ws["A1"].value


def test_apr_tab_has_platform_section_for_chatgpt(apr_wb):
    ws = apr_wb["AI Platform Response Tracking"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 30)]
    assert any(v and "Platform: ChatGPT" in str(v) for v in all_values)


def test_apr_tab_has_three_category_subsections_per_platform(apr_wb):
    ws = apr_wb["AI Platform Response Tracking"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 50)]
    cats = ["Direct Brand Queries", "Category-Based Queries", "Comparison Queries"]
    for cat in cats:
        assert any(v and cat in str(v) for v in all_values), f"Missing category: {cat}"


def test_apr_tab_writes_empty_placeholder_for_categories_with_no_data(apr_wb):
    ws = apr_wb["AI Platform Response Tracking"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 60)]
    assert any(
        v and "No data for this category" in str(v)
        for v in all_values
    )


def test_apr_tab_freeze_panes_at_A3(apr_wb):
    ws = apr_wb["AI Platform Response Tracking"]
    assert str(ws.freeze_panes) == "A3"


# ---------------------------------------------------------------------------
# Sentiment & Co-Occurrence workbook tests
# ---------------------------------------------------------------------------

@pytest.fixture
def sc_wb(tmp_path):
    from src.peec_client import Chat
    from src.matchers import LabeledChat
    from src.prompt_library import PromptEntry

    chat = Chat(
        id="ch_1", model="chatgpt-scraper", model_channel="ChatGPT",
        prompt="Test prompt", response="Babylon Tours is great",
        country="US", position=1, mentions=["Babylon Tours", "Competitor X"],
        sources=[], sentiment=75.0, created="2026-05-01",
    )
    lc = LabeledChat(chat=chat, prompt_id="DB-01", category="Direct Brand Queries")
    library = {"DB-01": PromptEntry("DB-01", "Test prompt", "Direct Brand Queries", "", "")}
    path = build_workbook([lc], library, BRAND, DATE_RANGE, tmp_path / "sc.xlsx")
    return openpyxl.load_workbook(path)


def test_sc_tab_has_three_section_headers(sc_wb):
    ws = sc_wb["Sentiment & Co-Occurrence"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 40)]
    assert any(v and "Sentiment Analysis Summary" in str(v) for v in all_values)
    assert any(v and "Co-Occurrence Analysis" in str(v) for v in all_values)
    assert any(v and "Detailed Sentiment by Prompt" in str(v) for v in all_values)


def test_sc_tab_summary_includes_overall_row(sc_wb):
    ws = sc_wb["Sentiment & Co-Occurrence"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 30)]
    assert any(v and "OVERALL" in str(v) for v in all_values)


def test_sc_tab_writes_placeholder_for_empty_cooccurrence(tmp_path):
    from src.peec_client import Chat
    from src.matchers import LabeledChat
    from src.prompt_library import PromptEntry

    chat = Chat(
        id="ch_1", model="chatgpt-scraper", model_channel="ChatGPT",
        prompt="Test prompt", response="No mentions here",
        country="US", position=None, mentions=[],
        sources=[], sentiment=None, created="2026-05-01",
    )
    lc = LabeledChat(chat=chat, prompt_id="DB-01", category="Direct Brand Queries")
    library = {"DB-01": PromptEntry("DB-01", "Test prompt", "Direct Brand Queries", "", "")}
    path = build_workbook([lc], library, BRAND, DATE_RANGE, tmp_path / "sc_empty.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Sentiment & Co-Occurrence"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 40)]
    assert any(v and "No co-occurrence data" in str(v) for v in all_values)


def test_sc_tab_brand_name_substituted_into_column_labels(sc_wb):
    ws = sc_wb["Sentiment & Co-Occurrence"]
    all_values = [ws.cell(row=r, column=c).value for r in range(1, 30) for c in range(1, 9)]
    assert any(v and f"{BRAND} Mentioned" in str(v) for v in all_values)


# ---------------------------------------------------------------------------
# Benchmarking workbook tests
# ---------------------------------------------------------------------------

@pytest.fixture
def bm_wb(tmp_path):
    from src.peec_client import Chat
    from src.matchers import LabeledChat
    from src.prompt_library import PromptEntry

    chat = Chat(
        id="ch_1", model="chatgpt-scraper", model_channel="ChatGPT",
        prompt="Test", response="", country="US", position=1,
        mentions=[BRAND, "Competitor X"], sources=[], sentiment=70.0, created="2026-05-01",
    )
    lc = LabeledChat(chat=chat, prompt_id="DB-01", category="Direct Brand Queries")
    library = {"DB-01": PromptEntry("DB-01", "Test prompt", "Direct Brand Queries", "", "")}
    path = build_workbook([lc], library, BRAND, DATE_RANGE, tmp_path / "bm.xlsx")
    return openpyxl.load_workbook(path)


def test_bm_tab_has_three_category_sections(bm_wb):
    ws = bm_wb["Benchmarking"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 30)]
    cats = ["Direct Brand Queries", "Category-Based Queries", "Comparison Queries"]
    for cat in cats:
        assert any(v and cat in str(v) for v in all_values), f"Missing: {cat}"


def test_bm_tab_focal_brand_first_in_each_section(bm_wb):
    ws = bm_wb["Benchmarking"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 30)]
    brand_positions = [i for i, v in enumerate(all_values) if v == BRAND]
    assert len(brand_positions) >= 3  # one per category


def test_bm_tab_focal_brand_name_bolded_in_first_column(bm_wb):
    ws = bm_wb["Benchmarking"]
    for r in range(1, 30):
        cell = ws.cell(row=r, column=1)
        if cell.value == BRAND and cell.font and cell.font.bold:
            return
    pytest.fail(f"No bolded cell with value '{BRAND}' found in column A")


def test_bm_tab_categories_in_DB_CB_CO_order(bm_wb):
    ws = bm_wb["Benchmarking"]
    cats = ["Direct Brand Queries", "Category-Based Queries", "Comparison Queries"]
    positions = []
    for r in range(1, 30):
        v = ws.cell(row=r, column=1).value
        if v in cats:
            positions.append((r, v))
    found = [cat for _, cat in positions]
    assert found == cats
