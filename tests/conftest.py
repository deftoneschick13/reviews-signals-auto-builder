"""Shared pytest fixtures."""

import json
from pathlib import Path

import openpyxl
import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture
def peec_response_json():
    with open(FIXTURES_DIR / "sample_peec_response.json") as f:
        return json.load(f)


@pytest.fixture
def sample_workbook_path(tmp_path):
    """Create a minimal Prompt Library workbook for testing."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Prompt Library")

    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "Tell me about Brand X", "Awareness", "High"])
    ws.append(["DB-02", "What is Brand X known for?", "Awareness", "High"])
    ws.append(["DB-03", "Brand X reviews", "Reviews", "Medium"])
    ws.append([])

    ws.append(["Category-Based Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["CB-01", "Best hotels in the area", "Discovery", "High"])
    ws.append(["CB-02", "Top resorts near me", "Discovery", "Medium"])
    ws.append(["CB-03", "Luxury hotels comparison", "Comparison", "Medium"])
    ws.append([])

    ws.append(["Comparison Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["CO-01", "Brand X vs Brand Y", "Comparison", "High"])
    ws.append(["CO-02", "Which hotel is better", "Comparison", "Medium"])
    ws.append(["CO-03", "Compare luxury resorts", "Comparison", "Low"])

    wb.create_sheet("Summary")  # unrelated extra tab

    path = tmp_path / "sample_prompt_library.xlsx"
    wb.save(path)
    return path
