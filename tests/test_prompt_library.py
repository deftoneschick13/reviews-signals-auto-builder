"""Tests for src/prompt_library.py."""

import openpyxl
import pytest

from src.prompt_library import PromptLibraryError, read_prompt_library


def test_returns_all_nine_entries_with_correct_categories(sample_workbook_path):
    library = read_prompt_library(sample_workbook_path)

    assert len(library) == 9

    for pid in ["DB-01", "DB-02", "DB-03"]:
        assert library[pid].category == "Direct Brand Queries"

    for pid in ["CB-01", "CB-02", "CB-03"]:
        assert library[pid].category == "Category-Based Queries"

    for pid in ["CO-01", "CO-02", "CO-03"]:
        assert library[pid].category == "Comparison Queries"

    assert library["DB-01"].text == "Tell me about Brand X"
    assert library["CB-01"].intent == "Discovery"
    assert library["CO-03"].priority == "Low"


def test_handles_case_insensitive_tab_name(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("prompt library")  # all lowercase
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "Some prompt", "", ""])
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    library = read_prompt_library(path)
    assert "DB-01" in library


def test_handles_tab_name_with_trailing_whitespace(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library ")  # trailing space
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "Some prompt", "", ""])
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    library = read_prompt_library(path)
    assert "DB-01" in library


def test_raises_when_tab_is_missing(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    with pytest.raises(PromptLibraryError, match="Prompt Library"):
        read_prompt_library(path)


def test_raises_on_duplicate_prompt_ids(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "First occurrence", "", ""])
    ws.append(["DB-01", "Duplicate", "", ""])
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    with pytest.raises(PromptLibraryError, match="Duplicate"):
        read_prompt_library(path)


def test_raises_when_prompt_id_has_no_text(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "", "", ""])  # empty text
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    with pytest.raises(PromptLibraryError, match="no text"):
        read_prompt_library(path)


def test_skips_blank_rows_between_sections(sample_workbook_path):
    library = read_prompt_library(sample_workbook_path)
    # Blank rows between sections should not produce entries or errors
    assert len(library) == 9


def test_skips_column_header_rows_within_sections(sample_workbook_path):
    library = read_prompt_library(sample_workbook_path)
    # "Prompt ID" should not appear as a key
    assert "Prompt ID" not in library
    assert "Prompt Text" not in library


def test_strips_whitespace_from_all_cell_values(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["  DB-01  ", "  Padded text  ", "  Intent  ", "  High  "])
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    library = read_prompt_library(path)
    assert "DB-01" in library
    assert library["DB-01"].text == "Padded text"
    assert library["DB-01"].intent == "Intent"
    assert library["DB-01"].priority == "High"


def test_raises_when_zero_entries_found(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    # No data rows
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    with pytest.raises(PromptLibraryError, match="No prompt entries"):
        read_prompt_library(path)


def test_handles_workbook_with_extra_unrelated_tabs(tmp_path):
    """Only the Prompt Library tab is read; other tabs are ignored."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Sales Data")
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "Tell me about Brand X", "Awareness", "High"])
    wb.create_sheet("Another Sheet")
    path = tmp_path / "extra_tabs.xlsx"
    wb.save(path)

    library = read_prompt_library(path)
    assert len(library) == 1
    assert "DB-01" in library


def test_handles_prompt_text_with_leading_trailing_whitespace(tmp_path):
    """Prompt text with surrounding whitespace is stripped on read."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "  Tell me about Brand X  ", "Awareness", "High"])
    path = tmp_path / "whitespace.xlsx"
    wb.save(path)

    library = read_prompt_library(path)
    assert library["DB-01"].text == "Tell me about Brand X"


def test_handles_intent_and_priority_as_empty_strings(tmp_path):
    """Rows with blank Intent and Priority are accepted and stored as ''."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Prompt Library")
    ws.append(["Direct Brand Queries"])
    ws.append(["Prompt ID", "Prompt Text", "Intent", "Priority"])
    ws.append(["DB-01", "Tell me about Brand X", None, None])
    path = tmp_path / "empty_intent.xlsx"
    wb.save(path)

    library = read_prompt_library(path)
    assert library["DB-01"].intent == ""
    assert library["DB-01"].priority == ""
